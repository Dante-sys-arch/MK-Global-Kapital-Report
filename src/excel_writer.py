"""
Generates the MK Global Kapital DACH Clipping Report as Excel file.
Matches the existing report format exactly.
"""
import json
from datetime import datetime
from pathlib import Path
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

DATA_FILE = Path(__file__).parent.parent / "data" / "clippings.json"
OUTPUT_DIR = Path(__file__).parent.parent / "output"


def load_clippings():
    if DATA_FILE.exists():
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return []


def build_report():
    clips = load_clippings()
    if not clips:
        print("No clippings to report")
        return None

    wb = Workbook()

    # ── Sheet 1: Clippings ──
    ws = wb.active
    ws.title = "2026 Clippings"

    # Header area (rows 1-7)
    ws.merge_cells("A1:G4")
    ws["A6"] = datetime.now()
    ws["A6"].number_format = "mm-dd-yy"
    ws["A6"].font = Font(name="Calibri", size=11)

    # Column headers (row 8)
    headers = ["Publication Date", "Media Outlet", "Title", "Country", "Type (Print / Online)", "Tier", "Link"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top")

    # Data rows
    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="top")
    date_align = Alignment(horizontal="left")

    for i, clip in enumerate(clips):
        row = 9 + i
        
        # Date
        cell_date = ws.cell(row=row, column=1)
        try:
            cell_date.value = datetime.strptime(clip.get("date", ""), "%Y-%m-%d")
        except (ValueError, TypeError):
            cell_date.value = clip.get("date", "")
        cell_date.number_format = "mm-dd-yy"
        cell_date.font = data_font
        cell_date.alignment = date_align

        # Other fields
        fields = [
            clip.get("outlet", ""),
            clip.get("title", ""),
            clip.get("country", ""),
            clip.get("type", ""),
            clip.get("tier", ""),
            clip.get("link", ""),
        ]
        for col, val in enumerate(fields, 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_align

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 70
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 8
    ws.column_dimensions["G"].width = 80

    # ── Sheet 2: Analysis ──
    wa = wb.create_sheet("Analysis 2026")

    title_font = Font(name="Calibri", size=14, bold=True)
    section_font = Font(name="Calibri", size=12, bold=True)
    label_font = Font(name="Calibri", size=11)
    value_font = Font(name="Calibri", size=11, bold=True)
    pct_fmt = "0.0%"

    # Title
    wa["A1"] = "Mikro Kapital Management — Quantitative Clippings Analysis (DACH) | 2026"
    wa["A1"].font = title_font
    wa["A2"] = f"Source: Sheet '2026 Clippings' (A9:G{8 + len(clips)})"
    wa["A2"].font = Font(name="Calibri", size=10, italic=True, color="666666")

    # Summary KPIs
    wa["A4"] = "Summary KPIs"
    wa["A4"].font = section_font
    
    dates = [c.get("date", "") for c in clips if c.get("date")]
    tier_counts = Counter(str(c.get("tier", "")) for c in clips if c.get("tier"))
    country_counts = Counter(c.get("country", "") for c in clips if c.get("country"))
    type_counts = Counter(c.get("type", "") for c in clips if c.get("type"))
    outlet_set = set(c.get("outlet", "") for c in clips if c.get("outlet"))

    kpis = [
        ("Total clippings", len(clips)),
        ("Period start (min date)", min(dates) if dates else ""),
        ("Period end (max date)", max(dates) if dates else ""),
        ("# Media outlets (unique)", len(outlet_set)),
        ("# Countries (unique)", len(country_counts)),
    ]
    for i, (label, val) in enumerate(kpis):
        wa.cell(row=5 + i, column=1, value=label).font = label_font
        cell = wa.cell(row=5 + i, column=2, value=val)
        cell.font = value_font
        if isinstance(val, str) and "-" in val:
            try:
                cell.value = datetime.strptime(val, "%Y-%m-%d")
                cell.number_format = "yyyy-mm-dd"
            except ValueError:
                pass

    # Tier analysis
    row_start = 10
    wa.cell(row=row_start, column=1, value="Tier-1 / Tier-2 Analysis").font = section_font
    wa.cell(row=row_start + 1, column=1, value="Tier").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=row_start + 1, column=2, value="Count").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=row_start + 1, column=3, value="Share").font = Font(name="Calibri", size=11, bold=True)

    total = len(clips) or 1
    for i, tier in enumerate(["1", "2"]):
        cnt = tier_counts.get(tier, 0) + tier_counts.get(int(tier) if tier.isdigit() else tier, 0)
        wa.cell(row=row_start + 2 + i, column=1, value=f"Tier {tier}").font = label_font
        wa.cell(row=row_start + 2 + i, column=2, value=cnt).font = label_font
        cell = wa.cell(row=row_start + 2 + i, column=3, value=cnt / total)
        cell.number_format = pct_fmt
        cell.font = label_font

    # Monthly breakdown
    wa.cell(row=row_start, column=5, value="Clippings by Month").font = section_font
    wa.cell(row=row_start + 1, column=5, value="Month").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=row_start + 1, column=6, value="Count").font = Font(name="Calibri", size=11, bold=True)

    month_counts = Counter()
    for c in clips:
        d = c.get("date", "")
        if d and len(d) >= 7:
            month_counts[d[:7]] += 1

    for i, (month, cnt) in enumerate(sorted(month_counts.items())):
        wa.cell(row=row_start + 2 + i, column=5, value=month).font = label_font
        wa.cell(row=row_start + 2 + i, column=6, value=cnt).font = label_font

    # Country breakdown
    cr = row_start + 6 + len(month_counts)
    wa.cell(row=cr, column=1, value="Clippings by Country").font = section_font
    wa.cell(row=cr + 1, column=1, value="Country").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=cr + 1, column=2, value="Count").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=cr + 1, column=3, value="Share").font = Font(name="Calibri", size=11, bold=True)

    for i, (country, cnt) in enumerate(sorted(country_counts.items())):
        wa.cell(row=cr + 2 + i, column=1, value=country).font = label_font
        wa.cell(row=cr + 2 + i, column=2, value=cnt).font = label_font
        cell = wa.cell(row=cr + 2 + i, column=3, value=cnt / total)
        cell.number_format = pct_fmt
        cell.font = label_font

    # Type breakdown
    wa.cell(row=cr, column=5, value="Clippings by Type").font = section_font
    wa.cell(row=cr + 1, column=5, value="Type").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=cr + 1, column=6, value="Count").font = Font(name="Calibri", size=11, bold=True)
    wa.cell(row=cr + 1, column=7, value="Share").font = Font(name="Calibri", size=11, bold=True)

    for i, (typ, cnt) in enumerate(sorted(type_counts.items())):
        wa.cell(row=cr + 2 + i, column=5, value=typ).font = label_font
        wa.cell(row=cr + 2 + i, column=6, value=cnt).font = label_font
        cell = wa.cell(row=cr + 2 + i, column=7, value=cnt / total)
        cell.number_format = pct_fmt
        cell.font = label_font

    # Column widths for analysis
    wa.column_dimensions["A"].width = 35
    wa.column_dimensions["B"].width = 15
    wa.column_dimensions["C"].width = 10
    wa.column_dimensions["E"].width = 20
    wa.column_dimensions["F"].width = 10
    wa.column_dimensions["G"].width = 10

    # Save
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    filename = f"{today}_MK_2026_DACH_Clipping_Report___Analysis.xlsx"
    filepath = OUTPUT_DIR / filename
    wb.save(filepath)
    print(f"Report saved: {filepath}")
    return filepath


if __name__ == "__main__":
    build_report()
