"""
Excel-Pressespiegel pro Kunde — Clippings-Sheet + Analyse-Sheet mit Charts.
Format entspricht dem in deutschen PR-Agenturen üblichen Pressespiegel.
"""
from collections import Counter
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .monitor import load_clippings


def build_report(cfg, clips=None):
    """Excel-Report für einen Kunden erzeugen. Gibt den Dateipfad zurück."""
    if clips is None:
        clips = load_clippings(cfg)
    if not clips:
        print(f"  Keine Clippings für {cfg.client_id} — kein Report")
        return None

    year = datetime.now().year
    wb = Workbook()

    # ── Sheet 1: Clippings ──
    ws = wb.active
    ws.title = f"{year} Clippings"

    ws.merge_cells("A1:G3")
    ws["A1"] = cfg.report_title
    ws["A1"].font = Font(name="Calibri", size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws["A5"] = datetime.now()
    ws["A5"].number_format = "dd.mm.yyyy"
    ws["A5"].font = Font(name="Calibri", size=11)

    headers = ["Datum", "Medium", "Titel", "Land", "Typ (Print/Online)", "Tier", "Link"]
    header_font = Font(name="Calibri", size=11, bold=True)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left", vertical="top")

    data_font = Font(name="Calibri", size=11)
    data_align = Alignment(horizontal="left", vertical="top")
    for i, clip in enumerate(clips):
        row = 8 + i
        cell_date = ws.cell(row=row, column=1)
        try:
            cell_date.value = datetime.strptime(clip.get("date", ""), "%Y-%m-%d")
        except (ValueError, TypeError):
            cell_date.value = clip.get("date", "")
        cell_date.number_format = "dd.mm.yy"
        cell_date.font = data_font
        cell_date.alignment = Alignment(horizontal="left")

        fields = [clip.get("outlet", ""), clip.get("title", ""), clip.get("country", ""),
                  clip.get("type", ""), clip.get("tier", ""), clip.get("link", "")]
        for col, val in enumerate(fields, 2):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = data_font
            cell.alignment = data_align

    for col, width in zip("ABCDEFG", [12, 30, 70, 8, 18, 8, 80]):
        ws.column_dimensions[col].width = width

    # ── Sheet 2: Analyse ──
    wa = wb.create_sheet(f"Analyse {year}")
    section_font = Font(name="Calibri", size=12, bold=True)
    label_font = Font(name="Calibri", size=11)
    value_font = Font(name="Calibri", size=11, bold=True)
    bold = Font(name="Calibri", size=11, bold=True)
    pct_fmt = "0.0%"

    wa["A1"] = f"{cfg.name} — Quantitative Clipping-Analyse (DACH) | {year}"
    wa["A1"].font = Font(name="Calibri", size=14, bold=True)

    dates = [c.get("date", "") for c in clips if c.get("date")]
    tier_counts = Counter(str(c.get("tier", "")) for c in clips if c.get("tier"))
    country_counts = Counter(c.get("country", "") for c in clips if c.get("country"))
    type_counts = Counter(c.get("type", "") for c in clips if c.get("type"))
    outlet_set = {c.get("outlet", "") for c in clips if c.get("outlet")}
    total = len(clips) or 1

    wa["A4"] = "Kennzahlen"
    wa["A4"].font = section_font
    kpis = [
        ("Clippings gesamt", len(clips)),
        ("Zeitraum von", min(dates) if dates else ""),
        ("Zeitraum bis", max(dates) if dates else ""),
        ("Medien (unique)", len(outlet_set)),
        ("Länder (unique)", len(country_counts)),
    ]
    for i, (label, val) in enumerate(kpis):
        wa.cell(row=5 + i, column=1, value=label).font = label_font
        wa.cell(row=5 + i, column=2, value=val).font = value_font

    # Tier-Verteilung
    row_start = 11
    wa.cell(row=row_start, column=1, value="Tier-Analyse").font = section_font
    for col, h in [(1, "Tier"), (2, "Anzahl"), (3, "Anteil")]:
        wa.cell(row=row_start + 1, column=col, value=h).font = bold
    for i, tier in enumerate(["1", "2"]):
        cnt = tier_counts.get(tier, 0)
        wa.cell(row=row_start + 2 + i, column=1, value=f"Tier {tier}").font = label_font
        wa.cell(row=row_start + 2 + i, column=2, value=cnt).font = label_font
        cell = wa.cell(row=row_start + 2 + i, column=3, value=cnt / total)
        cell.number_format = pct_fmt
        cell.font = label_font

    # Monatsverlauf
    wa.cell(row=row_start, column=5, value="Clippings pro Monat").font = section_font
    wa.cell(row=row_start + 1, column=5, value="Monat").font = bold
    wa.cell(row=row_start + 1, column=6, value="Anzahl").font = bold
    month_counts = Counter(c["date"][:7] for c in clips if len(c.get("date", "")) >= 7)
    for i, (month, cnt) in enumerate(sorted(month_counts.items())):
        wa.cell(row=row_start + 2 + i, column=5, value=month).font = label_font
        wa.cell(row=row_start + 2 + i, column=6, value=cnt).font = label_font

    # Länder
    cr = row_start + 6 + len(month_counts)
    wa.cell(row=cr, column=1, value="Clippings nach Land").font = section_font
    for col, h in [(1, "Land"), (2, "Anzahl"), (3, "Anteil")]:
        wa.cell(row=cr + 1, column=col, value=h).font = bold
    for i, (country, cnt) in enumerate(sorted(country_counts.items())):
        wa.cell(row=cr + 2 + i, column=1, value=country).font = label_font
        wa.cell(row=cr + 2 + i, column=2, value=cnt).font = label_font
        cell = wa.cell(row=cr + 2 + i, column=3, value=cnt / total)
        cell.number_format = pct_fmt
        cell.font = label_font

    # Typen
    wa.cell(row=cr, column=5, value="Clippings nach Typ").font = section_font
    for col, h in [(5, "Typ"), (6, "Anzahl"), (7, "Anteil")]:
        wa.cell(row=cr + 1, column=col, value=h).font = bold
    for i, (typ, cnt) in enumerate(sorted(type_counts.items())):
        wa.cell(row=cr + 2 + i, column=5, value=typ).font = label_font
        wa.cell(row=cr + 2 + i, column=6, value=cnt).font = label_font
        cell = wa.cell(row=cr + 2 + i, column=7, value=cnt / total)
        cell.number_format = pct_fmt
        cell.font = label_font

    for col, width in zip("ABCEFG", [35, 15, 10, 20, 10, 10]):
        wa.column_dimensions[col].width = width

    # ── Charts ──
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.chart.label import DataLabelList

    chart_row = cr + 2 + max(len(country_counts), len(type_counts)) + 4
    if month_counts:
        bc = BarChart()
        bc.type = "col"
        bc.title = "Clippings pro Monat"
        bc.style = 10
        data_ref = Reference(wa, min_col=6, min_row=row_start + 1,
                             max_row=row_start + 1 + len(month_counts), max_col=6)
        cats_ref = Reference(wa, min_col=5, min_row=row_start + 2,
                             max_row=row_start + 1 + len(month_counts))
        bc.add_data(data_ref, titles_from_data=True)
        bc.set_categories(cats_ref)
        bc.width = 18
        bc.height = 10
        bc.dataLabels = DataLabelList()
        bc.dataLabels.showVal = True
        wa.add_chart(bc, f"A{chart_row}")

    pc = PieChart()
    pc.title = "Tier-Verteilung"
    pc.style = 10
    tier_data = Reference(wa, min_col=2, min_row=row_start + 1, max_row=row_start + 3)
    tier_cats = Reference(wa, min_col=1, min_row=row_start + 2, max_row=row_start + 3)
    pc.add_data(tier_data, titles_from_data=True)
    pc.set_categories(tier_cats)
    pc.width = 12
    pc.height = 10
    pc.dataLabels = DataLabelList()
    pc.dataLabels.showPercent = True
    wa.add_chart(pc, f"E{chart_row}")

    # Speichern
    cfg.output_dir.mkdir(parents=True, exist_ok=True)
    today = datetime.now().strftime("%Y%m%d")
    filepath = cfg.output_dir / f"{today}_{cfg.client_id}_Pressespiegel.xlsx"
    wb.save(filepath)
    latest = cfg.output_dir / "latest_Pressespiegel.xlsx"
    wb.save(latest)
    print(f"  Report: {filepath}")
    return filepath
