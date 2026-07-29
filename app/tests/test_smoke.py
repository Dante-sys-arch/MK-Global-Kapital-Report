"""
Offline-Smoke-Tests: Config-Laden, Feed-Auflösung, Dedup, Report-Erzeugung.
Läuft ohne Netzwerk und ohne API-Keys:  python -m pytest app/tests/ -v
"""
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from engine.config import load_all_clients, load_client, ConfigError, CLIENTS_DIR
from engine.feeds import resolve_feeds, FEED_PACKS
from engine.monitor import is_duplicate, normalize_url, guess_country, save_clippings, load_clippings
from engine.report import build_report

SAMPLE_CLIPS = [
    {"date": "2026-07-01", "outlet": "Handelsblatt", "title": "Demo Capital legt neuen Fonds auf",
     "country": "D", "type": "Online", "tier": 1, "link": "https://www.handelsblatt.com/demo-1"},
    {"date": "2026-07-15", "outlet": "Finews", "title": "Erika Musterfrau im Interview über Impact Investing",
     "country": "CH", "type": "Online", "tier": 1, "link": "https://www.finews.ch/demo-2"},
    {"date": "2026-06-20", "outlet": "BondGuide", "title": "Demo Capital platziert Anleihe",
     "country": "D", "type": "Online", "tier": 2, "link": "https://www.bondguide.de/demo-3"},
]


def test_all_client_configs_valid():
    clients = load_all_clients()
    assert len(clients) >= 2, "Mindestens demo + mk-global-kapital erwartet"
    ids = [c.client_id for c in clients]
    assert "demo" in ids
    assert "mk-global-kapital" in ids
    for c in clients:
        assert c.keywords, f"{c.client_id}: keywords fehlen"


def test_template_is_ignored():
    clients = load_all_clients()
    assert all(c.client_id != "kundenkuerzel" for c in clients)


def test_invalid_config_raises(tmp_path):
    bad = tmp_path / "bad.yml"
    bad.write_text("name: Kaputt GmbH\nkeywords: []\n", encoding="utf-8")
    try:
        load_client(bad)
        raise AssertionError("ConfigError erwartet")
    except ConfigError:
        pass


def test_feed_packs_resolve():
    feeds = resolve_feeds(["finance_dach", "general_dach"], ["https://example.com/feed"])
    assert len(feeds) == len(set(feeds)), "Feeds müssen dedupliziert sein"
    assert "https://example.com/feed" in feeds
    assert len(feeds) > 40
    assert all(isinstance(pack, list) and pack for pack in FEED_PACKS.values())


def test_dedup_by_url_and_title():
    existing = SAMPLE_CLIPS
    dup_url = {"title": "Anderer Titel hier mit Laenge", "link": "https://www.handelsblatt.com/demo-1?utm_source=x"}
    dup_title = {"title": "Demo Capital legt neuen Fonds auf", "outlet": "Irgendwo", "link": "https://other.example/xyz"}
    fresh = {"title": "Voellig neuer Artikel ueber etwas anderes", "link": "https://neu.example/artikel"}
    assert is_duplicate(dup_url, existing)
    assert is_duplicate(dup_title, existing)
    assert not is_duplicate(fresh, existing)


def test_normalize_url_strips_tracking():
    assert normalize_url("https://X.de/a?utm_source=nl&utm_medium=mail#top") == "https://x.de/a"


def test_guess_country():
    assert guess_country("Finews", "https://www.finews.ch/x") == "CH"
    assert guess_country("Der Standard", "https://derstandard.at/x") == "A"
    assert guess_country("Handelsblatt", "https://handelsblatt.com/x") == "D"


def test_report_generation(tmp_path, monkeypatch):
    cfg = load_client(CLIENTS_DIR / "demo.yml")
    monkeypatch.setattr("engine.config.DATA_DIR", tmp_path / "data")
    monkeypatch.setattr("engine.config.OUTPUT_DIR", tmp_path / "output")
    save_clippings(cfg, SAMPLE_CLIPS)
    assert load_clippings(cfg) == SAMPLE_CLIPS

    path = build_report(cfg)
    assert path is not None and Path(path).exists()

    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert len(wb.sheetnames) == 2
    ws = wb[wb.sheetnames[0]]
    # Header in Zeile 7, Daten ab Zeile 8
    assert ws.cell(row=7, column=1).value == "Datum"
    assert ws.cell(row=8, column=3).value in [c["title"] for c in SAMPLE_CLIPS]


def test_digest_html():
    from engine.mailer import build_digest_html
    cfg = load_client(CLIENTS_DIR / "demo.yml")
    html = build_digest_html(cfg, SAMPLE_CLIPS[:2], 3)
    assert "Demo Capital AG" in html
    assert "2 neue Clippings" in html
    assert "handelsblatt.com" in html
